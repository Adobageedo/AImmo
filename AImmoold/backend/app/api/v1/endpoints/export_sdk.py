"""
Export SDK Endpoints - Génération d'exports Excel et PDF
Support exports de conversations, artefacts et données
"""

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import Response
from datetime import datetime, timedelta

from app.core.security import get_current_user_id
from app.core.supabase import get_supabase_client
from app.schemas.chat_sdk import (
    ExportRequest,
    ExportResponse,
    ExportFormat,
)
from app.services.export_service import (
    export_conversation_excel,
    export_conversation_pdf,
    export_table_excel,
    export_markdown,
    upload_to_storage,
)
from app.services.chat.chat_sdk_service import get_conversation_history


router = APIRouter()


# ============================================
# EXPORT CONVERSATIONS
# ============================================

@router.post("/conversation/excel")
async def export_conversation_to_excel(
    request: ExportRequest,
    user_id: str = Depends(get_current_user_id),
    supabase = Depends(get_supabase_client),
):
    """
    Exporte une conversation en Excel
    
    - Génère un fichier .xlsx avec messages et citations
    - Format tabulaire avec colonnes: Timestamp, Rôle, Message, Citations
    - Retourne le fichier directement ou une URL signée
    """
    if not request.conversation_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="conversation_id is required"
        )
    
        
    # Vérifier l'appartenance
    conv_response = supabase.table("conversations").select("user_id, title").eq(
        "id", request.conversation_id
    ).single().execute()
    
    if not conv_response.data or conv_response.data["user_id"] != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not your conversation"
        )
    
    # Récupérer les messages
    messages = await get_conversation_history(request.conversation_id, supabase)
    
    if not messages:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No messages to export"
        )
    
    # Générer l'Excel
    file_content, filename = await export_conversation_excel(
        conversation_id=request.conversation_id,
        messages=messages,
        include_citations=request.include_citations,
        supabase=supabase,
    )
    
    # Retourner directement le fichier
    return Response(
        content=file_content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(file_content)),
        }
    )


@router.post("/conversation/pdf")
async def export_conversation_to_pdf(
    request: ExportRequest,
    user_id: str = Depends(get_current_user_id),
    supabase = Depends(get_supabase_client),
):
    """
    Exporte une conversation en PDF
    
    - Génère un PDF formaté avec messages et citations
    - Design professionnel avec en-têtes et styles
    - Retourne le fichier directement
    """
    if not request.conversation_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="conversation_id is required"
        )
    
        
    # Vérifier l'appartenance
    conv_response = supabase.table("conversations").select("user_id, title").eq(
        "id", request.conversation_id
    ).single().execute()
    
    if not conv_response.data or conv_response.data["user_id"] != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not your conversation"
        )
    
    # Récupérer les messages
    messages = await get_conversation_history(request.conversation_id, supabase)
    
    if not messages:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No messages to export"
        )
    
    # Générer le PDF
    file_content, filename = await export_conversation_pdf(
        conversation_id=request.conversation_id,
        messages=messages,
        include_citations=request.include_citations,
        supabase=supabase,
    )
    
    # Retourner directement le fichier
    return Response(
        content=file_content,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(file_content)),
        }
    )


# ============================================
# EXPORT ARTEFACTS
# ============================================

@router.post("/artifact/{artifact_id}/excel")
async def export_artifact_to_excel(
    artifact_id: str,
    user_id: str = Depends(get_current_user_id),
    supabase = Depends(get_supabase_client),
):
    """
    Exporte un artefact de type table en Excel
    
    - Récupère l'artefact
    - Extrait les données tabulaires
    - Génère un fichier Excel formaté
    """
        
    # Récupérer l'artefact
    artifact_response = supabase.table("artifacts").select("*").eq(
        "id", artifact_id
    ).eq("user_id", user_id).single().execute()
    
    if not artifact_response.data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Artifact not found"
        )
    
    artifact = artifact_response.data
    
    # Vérifier que c'est un tableau
    if artifact["type"] != "table":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only table artifacts can be exported to Excel"
        )
    
    # Extraire les données
    content = artifact["content"]
    headers = content.get("headers", [])
    rows = content.get("rows", [])
    
    if not headers or not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Table has no data to export"
        )
    
    # Générer l'Excel
    file_content, filename = await export_table_excel(
        headers=headers,
        rows=rows,
        title=artifact.get("title", "Table"),
    )
    
    # Retourner le fichier
    return Response(
        content=file_content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(file_content)),
        }
    )


@router.post("/artifacts/bulk/excel")
async def export_multiple_artifacts_to_excel(
    artifact_ids: list[str],
    user_id: str = Depends(get_current_user_id),
    supabase = Depends(get_supabase_client),
):
    """
    Exporte plusieurs artefacts dans un seul fichier Excel
    
    - Crée un classeur avec une feuille par artefact
    - Supporte uniquement les artefacts de type table
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    import io
    
        
    # Créer le workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut
    
    # Pour chaque artefact
    for artifact_id in artifact_ids:
        # Récupérer l'artefact
        artifact_response = supabase.table("artifacts").select("*").eq(
            "id", artifact_id
        ).eq("user_id", user_id).single().execute()
        
        if not artifact_response.data or artifact_response.data["type"] != "table":
            continue
        
        artifact = artifact_response.data
        content = artifact["content"]
        headers = content.get("headers", [])
        rows = content.get("rows", [])
        
        if not headers or not rows:
            continue
        
        # Créer une feuille
        title = artifact.get("title", f"Table {artifact_id[:8]}")[:31]
        ws = wb.create_sheet(title=title)
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # En-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Données
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    if not wb.worksheets:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No valid table artifacts found"
        )
    
    # Sauvegarder
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"artifacts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(buffer.getvalue())),
        }
    )


# ============================================
# EXPORT MARKDOWN
# ============================================

@router.post("/markdown")
async def export_markdown_content(
    request: ExportRequest,
    user_id: str = Depends(get_current_user_id),
):
    """
    Exporte du contenu Markdown en HTML
    
    - Convertit le Markdown en HTML formaté
    - Template professionnel avec styles
    """
    if not request.content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="content is required"
        )
    
    # Générer le HTML
    file_content, filename = await export_markdown(request.content)
    
    # Retourner le fichier
    return Response(
        content=file_content,
        media_type="text/html",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(file_content)),
        }
    )
