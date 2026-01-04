"""
Export Service - Génération d'exports Excel et PDF
Support des exports de conversations, artefacts et données
"""

from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta
import io
import os
import tempfile
from uuid import uuid4

# Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# PDF
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors

# Markdown
import markdown

from app.schemas.chat_sdk import ExportFormat, Message, Citation
from app.core.config import settings


async def export_conversation_excel(
    conversation_id: str,
    messages: List[Message],
    include_citations: bool = True,
    supabase = None,
) -> tuple[bytes, str]:
    """
    Exporte une conversation en Excel
    """
    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conversation"
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # En-têtes
    headers = ["Timestamp", "Rôle", "Message"]
    if include_citations:
        headers.append("Citations")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Données
    for row_idx, message in enumerate(messages, 2):
        ws.cell(row=row_idx, column=1, value=message.created_at.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_idx, column=2, value=message.role.value.upper())
        ws.cell(row=row_idx, column=3, value=message.content)
        
        if include_citations and message.citations:
            citations_text = "\n".join([
                f"• {c.document_title} (Score: {c.relevance_score:.0%})"
                for c in message.citations
            ])
            ws.cell(row=row_idx, column=4, value=citations_text)
    
    # Ajuster les largeurs de colonnes
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 80
    if include_citations:
        ws.column_dimensions['D'].width = 40
    
    # Ajouter des bordures
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(messages) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # Sauvegarder dans un buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"conversation_{conversation_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return buffer.getvalue(), filename


async def export_conversation_pdf(
    conversation_id: str,
    messages: List[Message],
    include_citations: bool = True,
    supabase = None,
) -> tuple[bytes, str]:
    """
    Exporte une conversation en PDF
    """
    buffer = io.BytesIO()
    
    # Créer le document
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=18,
    )
    
    # Styles
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
    )
    
    user_style = ParagraphStyle(
        'UserMessage',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=10,
        leftIndent=20,
    )
    
    assistant_style = ParagraphStyle(
        'AssistantMessage',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#2c5282'),
        spaceAfter=10,
        leftIndent=20,
    )
    
    citation_style = ParagraphStyle(
        'Citation',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#718096'),
        leftIndent=40,
        spaceAfter=5,
    )
    
    # Contenu
    story = []
    
    # Titre
    story.append(Paragraph("Conversation Export", title_style))
    story.append(Spacer(1, 12))
    
    # Date d'export
    export_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    story.append(Paragraph(f"Exporté le: {export_date}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Messages
    for message in messages:
        # En-tête du message
        timestamp = message.created_at.strftime("%d/%m/%Y %H:%M:%S")
        role = message.role.value.upper()
        
        header = f"<b>{role}</b> - {timestamp}"
        story.append(Paragraph(header, styles['Heading3']))
        story.append(Spacer(1, 6))
        
        # Contenu
        style = user_style if message.role.value == "user" else assistant_style
        
        # Échapper le HTML et convertir les retours à la ligne
        content = message.content.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        content = content.replace('\n', '<br/>')
        
        story.append(Paragraph(content, style))
        
        # Citations
        if include_citations and message.citations:
            story.append(Spacer(1, 6))
            story.append(Paragraph("<b>Sources:</b>", styles['Heading4']))
            
            for citation in message.citations:
                citation_text = f"• {citation.document_title} (Score: {citation.relevance_score:.0%})"
                story.append(Paragraph(citation_text, citation_style))
        
        story.append(Spacer(1, 20))
    
    # Générer le PDF
    doc.build(story)
    buffer.seek(0)
    
    filename = f"conversation_{conversation_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return buffer.getvalue(), filename


async def export_table_excel(
    headers: List[str],
    rows: List[List[Any]],
    title: str = "Export",
) -> tuple[bytes, str]:
    """
    Exporte un tableau en Excel
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Max 31 caractères pour un nom de feuille
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # En-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Données
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Ajuster les largeurs
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Bordures
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(rows) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # Sauvegarder
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"{title.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return buffer.getvalue(), filename


async def export_markdown(
    content: str,
) -> tuple[bytes, str]:
    """
    Exporte du contenu Markdown
    """
    # Convertir en HTML
    html_content = markdown.markdown(
        content,
        extensions=['tables', 'fenced_code', 'codehilite']
    )
    
    # Template HTML complet
    full_html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Export</title>
        <style>
            body {{
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                line-height: 1.6;
                max-width: 800px;
                margin: 0 auto;
                padding: 20px;
                color: #333;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
                margin: 20px 0;
            }}
            th, td {{
                border: 1px solid #ddd;
                padding: 12px;
                text-align: left;
            }}
            th {{
                background-color: #366092;
                color: white;
            }}
            code {{
                background-color: #f4f4f4;
                padding: 2px 6px;
                border-radius: 3px;
            }}
            pre {{
                background-color: #f4f4f4;
                padding: 15px;
                border-radius: 5px;
                overflow-x: auto;
            }}
        </style>
    </head>
    <body>
        {html_content}
    </body>
    </html>
    """
    
    filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
    
    return full_html.encode('utf-8'), filename


async def upload_to_storage(
    file_content: bytes,
    filename: str,
    bucket: str = "exports",
    supabase = None,
) -> str:
    """
    Upload un fichier vers Supabase Storage et retourne l'URL
    """
    # Upload vers Supabase Storage
    file_path = f"{datetime.now().strftime('%Y/%m/%d')}/{uuid4()}_{filename}"
    
    response = supabase.storage.from_(bucket).upload(
        file_path,
        file_content,
        {
            "content-type": get_content_type(filename),
            "cache-control": "3600",
        }
    )
    
    # Générer une URL signée (expire dans 1 heure)
    signed_url = supabase.storage.from_(bucket).create_signed_url(
        file_path,
        3600,  # 1 heure
    )
    
    return signed_url['signedURL']


def get_content_type(filename: str) -> str:
    """
    Détermine le content-type d'un fichier
    """
    ext = filename.split('.')[-1].lower()
    
    content_types = {
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'pdf': 'application/pdf',
        'html': 'text/html',
        'md': 'text/markdown',
        'json': 'application/json',
    }
    
    return content_types.get(ext, 'application/octet-stream')
